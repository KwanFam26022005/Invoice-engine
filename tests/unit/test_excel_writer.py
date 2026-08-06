"""Unit tests for ExcelWriter verification."""

from pathlib import Path
import tempfile
import openpyxl

from document_benchmark.export.excel_writer import ExcelWriter


def test_excel_writer_benchmark_report():
    with tempfile.TemporaryDirectory() as tmp_dir:
        out_path = Path(tmp_dir) / "benchmark_report_test.xlsx"
        writer = ExcelWriter()

        run_info = {"run_id": "run_test_123", "timestamp": "2026-08-06T09:00:00Z"}
        documents = [{"document_id": "doc1", "filename": "sample.pdf", "page_count": 1, "sha256": "hash", "document_family": "invoice"}]
        engine_summary = [{"config_id": "mock_default", "doc_count": 1, "total_subtotal": 10000.0, "total_vat": 1000.0, "total_amount": 11000.0}]
        field_comparisons = [{"document_id": "doc1", "field_path": "total_amount", "consensus_value": "11000.0", "agreement_count": 1, "total_engines": 1, "disagreement_severity": "INFO"}]
        latency_metrics = [{"config_id": "mock_default", "run_count": 1, "success_count": 1, "mean_latency_ms": 100.0, "p50_latency_ms": 100.0, "p95_latency_ms": 100.0, "peak_ram_mb": 50.0}]
        validation_issues = []

        writer.write_benchmark_report(
            output_path=out_path,
            run_info=run_info,
            documents=documents,
            engine_summary=engine_summary,
            field_comparisons=field_comparisons,
            latency_metrics=latency_metrics,
            validation_issues=validation_issues,
        )

        assert out_path.exists()

        # Open with openpyxl and verify sheets and values
        wb = openpyxl.load_workbook(out_path)
        sheet_names = wb.sheetnames
        assert "RunInfo" in sheet_names
        assert "Documents" in sheet_names
        assert "EngineSummary" in sheet_names
        assert "FieldComparison" in sheet_names
        assert "LatencyMetrics" in sheet_names

        ws_sum = wb["EngineSummary"]
        assert ws_sum.cell(row=2, column=1).value == "mock_default"
        assert ws_sum.cell(row=2, column=5).value == 11000.0
