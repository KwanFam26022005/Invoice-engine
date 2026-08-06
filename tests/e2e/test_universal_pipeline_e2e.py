"""End-to-end synthetic integration tests for Universal Document Engine."""

from decimal import Decimal
from pathlib import Path
import fitz
import openpyxl
import pytest

from document_engine.export.exporter import ExcelExporter
from document_engine.orchestration.pipeline import DocumentPipeline
from document_engine.settings import AppConfig, get_workspace_paths
from document_engine.storage.database import DuckDBStorage


def create_synthetic_sales_invoice_pdf(path: Path) -> Path:
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    page.insert_text((50, 50), "HOA DON GIA TRI GIA TANG\nSo: INV-2026-888\nNgay 15/08/2026")
    page.insert_text((50, 120), "Ma so thue ben ban: 0101234567\nMa so thue ben mua: 0109876543")
    page.insert_text((50, 200), "Dich vu tu van giai phap IT\nTong tien thanh toan: 12.500.000 VND")
    doc.save(path)
    doc.close()
    return path


def create_synthetic_tax_cert_pdf(path: Path) -> Path:
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    page.insert_text((50, 50), "CHUNG TU KHAU TRU THUE THU NHAP CA NHAN\nSo: TNCN-2026-001")
    page.insert_text((50, 120), "So thue thu nhap ca nhan da khau tru: 1.500.000 VND")
    doc.save(path)
    doc.close()
    return path


def test_synthetic_sales_invoice_e2e(tmp_path: Path):
    ws_dir = tmp_path / "workspace"
    pdf_path = create_synthetic_sales_invoice_pdf(tmp_path / "sales_inv.pdf")

    config = AppConfig(workspace_root=str(ws_dir))
    pipeline = DocumentPipeline(config=config)

    res = pipeline.process_file(pdf_path)

    assert res.document_id.startswith("doc_")
    assert res.pdf_profile == "native_pdf"
    assert res.selected_parser == "pymupdf_native"
    assert res.document_family == "sales_invoice"
    assert res.requires_review is False

    # Check Excel Exporter
    paths = get_workspace_paths(ws_dir)
    storage = DuckDBStorage(paths.database_file)
    exporter = ExcelExporter(storage)

    excel_file = paths.exports / "document_export_e2e.xlsx"
    out_path = exporter.export_run("e2e_run", excel_file)

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert "Invoices" in wb.sheetnames
    assert "Validation" in wb.sheetnames


def test_synthetic_tax_certificate_e2e(tmp_path: Path):
    ws_dir = tmp_path / "workspace"
    pdf_path = create_synthetic_tax_cert_pdf(tmp_path / "tax_cert.pdf")

    config = AppConfig(workspace_root=str(ws_dir))
    pipeline = DocumentPipeline(config=config)

    res = pipeline.process_file(pdf_path)

    assert res.document_family == "tax_withholding_certificate"
    assert res.validation_status == "accepted"


def test_batch_processing_continue_on_error(tmp_path: Path):
    ws_dir = tmp_path / "workspace"
    folder = tmp_path / "inbox"
    folder.mkdir()

    create_synthetic_sales_invoice_pdf(folder / "doc1.pdf")
    create_synthetic_tax_cert_pdf(folder / "doc2.pdf")

    config = AppConfig(workspace_root=str(ws_dir))
    pipeline = DocumentPipeline(config=config)

    summary, results = pipeline.process_folder(folder)

    assert summary.received == 2
    assert summary.processed == 2
    assert len(results) == 2
