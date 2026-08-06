"""Excel report writer generating professional benchmark_report.xlsx and selected_business_output.xlsx."""

from pathlib import Path
from typing import Any, Dict, List
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """Generates formatted Excel workbooks for benchmark reports and business outputs."""

    def __init__(self) -> None:
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.border_thin = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

    def write_benchmark_report(
        self,
        output_path: Path,
        run_info: Dict[str, Any],
        documents: List[Dict[str, Any]],
        engine_summary: List[Dict[str, Any]],
        field_comparisons: List[Dict[str, Any]],
        latency_metrics: List[Dict[str, Any]],
        validation_issues: List[Dict[str, Any]],
    ) -> Path:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # 1. RunInfo
        ws_info = wb.create_sheet(title="RunInfo")
        ws_info.append(["Property", "Value"])
        for k, v in run_info.items():
            ws_info.append([str(k), str(v)])

        # 2. Documents
        ws_docs = wb.create_sheet(title="Documents")
        ws_docs.append(["Document ID", "Filename", "Page Count", "SHA256", "Document Family"])
        for d in documents:
            ws_docs.append([d.get("document_id"), d.get("filename"), d.get("page_count"), d.get("sha256"), d.get("document_family")])

        # 3. EngineSummary
        ws_summary = wb.create_sheet(title="EngineSummary")
        ws_summary.append(["Config ID", "Doc Count", "Subtotal Total", "VAT Total", "Total Amount Total"])
        for s in engine_summary:
            ws_summary.append([
                s.get("config_id"),
                s.get("doc_count"),
                float(s.get("total_subtotal") or 0.0),
                float(s.get("total_vat") or 0.0),
                float(s.get("total_amount") or 0.0),
            ])

        # 4. FieldComparison
        ws_fields = wb.create_sheet(title="FieldComparison")
        ws_fields.append(["Document ID", "Field Path", "Consensus Value", "Agreement Count", "Total Engines", "Disagreement Severity"])
        for f in field_comparisons:
            ws_fields.append([
                f.get("document_id"),
                f.get("field_path"),
                str(f.get("consensus_value")),
                f.get("agreement_count"),
                f.get("total_engines"),
                str(f.get("disagreement_severity")),
            ])

        # 5. LatencyMetrics
        ws_lat = wb.create_sheet(title="LatencyMetrics")
        ws_lat.append(["Config ID", "Run Count", "Success Count", "Mean Latency (ms)", "P50 (ms)", "P95 (ms)", "Peak RAM (MB)"])
        for m in latency_metrics:
            ws_lat.append([
                m.get("config_id"),
                m.get("run_count"),
                m.get("success_count"),
                round(m.get("mean_latency_ms") or 0.0, 2),
                round(m.get("p50_latency_ms") or 0.0, 2),
                round(m.get("p95_latency_ms") or 0.0, 2),
                round(m.get("peak_ram_mb") or 0.0, 2),
            ])

        # 6. ValidationIssues
        ws_val = wb.create_sheet(title="ValidationIssues")
        ws_val.append(["Engine", "Document ID", "Code", "Severity", "Field Path", "Message"])
        for issue in validation_issues:
            ws_val.append([
                issue.get("source_engine"),
                issue.get("document_id"),
                issue.get("code"),
                issue.get("severity"),
                issue.get("field_path"),
                issue.get("message"),
            ])

        # Apply formatting across all sheets
        for ws in wb.worksheets:
            self._format_worksheet(ws)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def write_business_output(
        self,
        output_path: Path,
        invoices: List[Dict[str, Any]],
        office_requests: List[Dict[str, Any]],
        software_proposals: List[Dict[str, Any]],
    ) -> Path:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # 1. Summary
        ws_sum = wb.create_sheet(title="Summary")
        ws_sum.append(["Category", "Count"])
        ws_sum.append(["Invoices", len(invoices)])
        ws_sum.append(["Office Supply Requests", len(office_requests)])
        ws_sum.append(["Software Proposals", len(software_proposals)])

        # 2. AllInvoices
        ws_inv = wb.create_sheet(title="AllInvoices")
        ws_inv.append([
            "Invoice Number",
            "Series",
            "Invoice Date",
            "Seller Tax ID",
            "Seller Name",
            "Buyer Tax ID",
            "Buyer Name",
            "Subtotal",
            "VAT",
            "Total Amount",
        ])
        for inv in invoices:
            p = inv.get("canonical_payload", {})
            ws_inv.append([
                p.get("invoice_number"),
                p.get("invoice_series"),
                p.get("invoice_date"),
                p.get("seller_tax_id"),
                p.get("seller_name"),
                p.get("buyer_tax_id"),
                p.get("buyer_name"),
                float(p.get("subtotal") or 0.0),
                float(p.get("vat_amount") or 0.0),
                float(p.get("total_amount") or 0.0),
            ])

        for ws in wb.worksheets:
            self._format_worksheet(ws)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def _format_worksheet(self, ws) -> None:
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        # Format header row
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))
                cell.border = self.border_thin
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
