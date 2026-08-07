"""Multi-sheet Excel Exporter using openpyxl with formula injection sanitization."""

from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from document_engine.storage.database import DuckDBStorage


class ExcelExporter:
    def __init__(self, db_storage: DuckDBStorage):
        self.db = db_storage

    def sanitize_cell_value(self, val: Any) -> Any:
        """Sanitize string values starting with '=', '+', '-', '@' against formula injection."""
        if isinstance(val, str) and val.startswith(("=", "+", "-", "@")):
            return f"'{val}"
        return val

    def export_run(self, run_id: str, output_path: Path) -> Path:
        output_path = Path(output_path).resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        sheet_configs = [
            ("Documents", "SELECT document_id, filename, status, received_at FROM documents;"),
            ("Invoices", "SELECT document_id, document_family, document_number, issue_date, currency, grand_total FROM business_documents;"),
            ("Parties", "SELECT party_id, document_id, role, name, tax_id, address FROM parties;"),
            ("LineItems", "SELECT line_item_id, document_id, description, quantity, unit, unit_price, amount FROM line_items;"),
            ("UsageRecords", "SELECT * FROM meter_readings LIMIT 0;"),
            ("MeterReadings", "SELECT reading_id, document_id, meter_number, opening_reading, closing_reading, consumption FROM meter_readings;"),
            ("PricingTiers", "SELECT * FROM line_items LIMIT 0;"),
            ("ServiceVolumes", "SELECT * FROM line_items LIMIT 0;"),
            ("PortServices", "SELECT container_id, document_id, container_number, container_size, teu, amount FROM container_records;"),
            ("TaxCertificates", "SELECT certificate_id, document_id, certificate_number, taxable_income, withheld_tax FROM tax_certificates;"),
            ("SupportingStatements", "SELECT * FROM line_items LIMIT 0;"),
            ("Validation", "SELECT issue_id, document_id, code, severity, field_path, message FROM validation_issues;"),
            ("Provenance", "SELECT attempt_id, document_id, parser_id, execution_time_seconds FROM parser_attempts;"),
            ("ReviewQueue", "SELECT review_id, document_id, reason, status, created_at FROM review_queue;"),
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        with self.db.get_connection() as conn:
            for sheet_name, query in sheet_configs:
                ws = wb.create_sheet(title=sheet_name)
                try:
                    rel = conn.execute(query)
                    headers = [desc[0] for desc in rel.description] if rel.description else ["id"]
                    rows = rel.fetchall()

                    # Write header
                    ws.append(headers)
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill

                    # Write data rows
                    for row in rows:
                        sanitized_row = [self.sanitize_cell_value(val) for val in row]
                        ws.append(sanitized_row)

                    # Formatting: freeze pane & auto-filter
                    ws.freeze_panes = "A2"
                    if len(rows) > 0:
                        ws.auto_filter.ref = ws.dimensions

                    # Adjust column widths
                    for col in ws.columns:
                        max_len = max(len(str(cell.value or "")) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                except Exception:
                    ws.append(["status"])
                    ws.append(["no data"])

        wb.save(output_path)
        return output_path
